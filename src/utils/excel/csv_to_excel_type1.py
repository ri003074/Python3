import pandas
from openpyxl import load_workbook
from openpyxl.chart import LineChart
from openpyxl.chart import Reference


data_path = "./data/"


class Excel:
    """expected csv format is following
    pin, test1, test2,,,
    p1, 1, 2,,
    p2, 3, 4,,
    p3, 5, 6,,
    """

    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        self.max_col = 0
        self.max_row = 0
        self.min_col = 0
        self.min_row = 0

    def make_excel_file(self):
        df = pandas.read_csv(self.input_file)
        df = df.loc[:, ~df.columns.str.match("Unnamed")]
        df.to_excel(self.output_file, index=False)
        self.max_col = len(df.columns)
        self.max_row = len(df.index) + 1

    def make_graph(
        self,
        y_axis_min,
        y_axis_max,
        x_axis_title=None,
        y_axis_title=None,
        graph_title=None,
        height=10,
        width=20,
        legend=False,
        legend_position="b",
        marker_symbol=None,
        line_no_fill=False,
    ):
        wb = load_workbook(self.output_file)
        ws = wb.active

        values = Reference(
            ws, min_col=2, min_row=1, max_col=self.max_col, max_row=self.max_row
        )
        categories = Reference(
            ws, min_col=1, min_row=2, max_col=1, max_row=self.max_row
        )
        chart = LineChart()

        if legend:
            chart.legend.position = legend_position
        else:
            chart.legend = None

        chart.height = height
        chart.width = width
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        chart.y_axis.scaling.min = y_axis_min
        chart.y_axis.scaling.max = y_axis_max
        chart.title = graph_title
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)

        for i in range(len(chart.series)):
            series = chart.series[i]
            if marker_symbol:
                series.marker.symbol = marker_symbol

            series.graphicalProperties.line.noFill = line_no_fill

        ws.add_chart(chart, "B4")
        wb.save(self.output_file)


if __name__ == "__main__":

    excel1 = Excel(data_path + "csv_type1_1.csv", data_path + "csv_type1_1.xlsx")
    excel1.make_excel_file()
    excel1.make_graph(
        graph_title="sample1",
        x_axis_title="abc",
        legend=True,
        y_axis_min=0,
        y_axis_max=10,
    )

    excel2 = Excel(data_path + "csv_type1_2.csv", data_path + "csv_type1_2.xlsx")
    excel2.make_excel_file()
    excel2.make_graph(
        y_axis_min=0, y_axis_max=10, marker_symbol="circle", line_no_fill=True
    )
