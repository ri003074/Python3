import csv
import datetime
import os
import re
import sys

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import win32com.client

from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# from glob import glob

image_count = 0

now = datetime.datetime.now()
date_now = now.strftime("%Y%m%d%H%M")


class WaveData:
    def __init__(
        self,
        file_name,
        folder_path,
        new_presentation,
        groupby="",
        header=[],
        index="Pin_Rate",
    ):
        self.data_df = ""
        self.file_name = file_name
        self.folder_path = folder_path
        self.groupby = groupby
        self.header = header
        self.index = index
        self.input_file = self.folder_path + self.file_name

        if new_presentation:
            pptx = win32com.client.Dispatch("PowerPoint.Application")
            pptx.Visible = True
            self.active_presentation = pptx.Presentations.Open(
                os.getcwd() + "/advtemplate_mini.pptx"
            )
        else:
            pptx = win32com.client.GetActiveObject("PowerPoint.Application")
            self.active_presentation = pptx.ActivePresentation

        self.slide_width = self.active_presentation.PageSetup.SlideWidth
        self.slide_height = self.active_presentation.PageSetup.SlideHeight
        self.slide_count = self.active_presentation.Slides.Count
        self.make_df_and_xlsx()

    def make_df_and_xlsx(self):
        """Make pandas dataframe and xlsx data from csv data

        Args:
            None

        Returns:
            None
        """
        with open(self.input_file, mode="r", encoding="utf-8-sig") as csvfile:
            reader = csv.reader(csvfile)
            data = []

            for rows in reader:
                match = re.match(r"(P(\d*).*?)_.*?_(.*?_.*?_.*?)_(.*?)_", rows[0])
                if match:
                    rows.insert(0, "Condition")
                    rows.insert(2, "Pin")
                    rows.insert(3, match.group(1))
                    rows.insert(4, "Pkind")
                    pin_num = int(match.group(2))
                    if pin_num < 1857:
                        pin_kind = "IO"
                    elif pin_num >= 1857 and pin_num <= 1888:
                        pin_kind = "WCK"
                    elif pin_num >= 1889 and pin_num <= 1890:
                        pin_kind = "CK"
                    elif pin_num >= 1921 and pin_num <= 1933:
                        pin_kind = "CA"
                    elif pin_num >= 1953 and pin_num <= 1959:
                        pin_kind = "CS"
                    else:
                        print("Pkind Error")
                        sys.exit()

                    rows.insert(5, pin_kind)
                    rows.insert(6, "Vi")
                    rows.insert(7, match.group(3).replace("00V", "V"))
                    rows.insert(8, "Rate")
                    rows.insert(
                        9, match.group(4).replace("Rate0r", "").replace("ns", "ps")
                    )

                dic = OrderedDict()
                for i in range(0, len(rows) - 1, 2):
                    try:
                        dic[rows[i].replace(" ", "").capitalize()] = float(
                            rows[i + 1].replace(" ", "")
                        )
                    except ValueError:
                        dic[rows[i].replace(" ", "").capitalize()] = rows[
                            i + 1
                        ].replace(" ", "")

                data.append(dic)

            self.data_df = pd.DataFrame(data)

            self.data_df.insert(
                1, "Pin_Rate", self.data_df["Pin"] + "_" + self.data_df["Rate"]
            )
            self.data_df.insert(
                2, "Pin_Vi", self.data_df["Pin"] + "_" + self.data_df["Vi"]
            )
            self.data_df.insert(
                3, "Pkind_Vi", self.data_df["Pkind"] + "_" + self.data_df["Vi"]
            )

            # adjust unit of dataframe
            self.adjust_unit()

            if self.header:
                self.data_df = self.data_df.set_axis(self.header, axis="columns")

            self.data_df = self.data_df.set_index(self.index)

            with pd.ExcelWriter(self.input_file.replace("csv", "xlsx")) as writer:
                self.data_df.to_excel(writer, sheet_name="summary")

                if self.groupby:
                    for name, group in self.data_df.groupby(self.groupby):
                        group.to_excel(writer, sheet_name=name)

    def make_excel_graph(
        self,
        chart_height=9,
        chart_width=16,
        chart_yaxis_title="",
        chart_yaxis_scaling_min="",
        chart_yaxis_scaling_max="",
        chart_yaxis_major_unit="",
        chart_position="E5",
        data_start_column=0,
        data_end_column=0,
    ):
        """ make specified excel graph using xlsx data

        """

        wb = load_workbook(self.input_file.replace("csv", "xlsx"))
        for i in range(len(wb.worksheets)):
            ws = wb.worksheets[i]

            values = Reference(
                ws,
                min_row=1,
                min_col=data_start_column,
                max_row=ws.max_row,
                max_col=data_end_column,
            )
            categories = Reference(
                ws, min_row=2, min_col=1, max_row=ws.max_row, max_col=1
            )

            self.setup_excel_chart(
                values=values,
                categories=categories,
                chart_height=chart_height,
                chart_width=chart_width,
                chart_yaxis_titles=chart_yaxis_title,
                chart_yaxis_scaling_mins=chart_yaxis_scaling_min,
                chart_yaxis_scaling_maxes=chart_yaxis_scaling_max,
                chart_yaxis_major_unit=chart_yaxis_major_unit,
            )

            ws.add_chart(self.chart, chart_position)
        wb.save(self.input_file.replace("csv", "xlsx"))

    def make_excel_graphs(
        self,
        chart_height=9,
        chart_width=16,
        chart_yaxis_titles=[],
        chart_yaxis_scaling_mins=[],
        chart_yaxis_scaling_maxes=[],
        chart_yaxis_major_unit=[],
        data_start_column=0,
    ):
        """ make excel graphs using xlsx data

        """

        wb = load_workbook(self.input_file.replace("csv", "xlsx"))
        for i in range(len(wb.worksheets)):
            ws = wb.worksheets[i]

            for i in range(ws.max_column + data_start_column * -1 + 1):
                values = Reference(
                    ws,
                    min_row=1,
                    min_col=data_start_column + i,
                    max_row=ws.max_row,
                    max_col=data_start_column + i,
                )
                categories = Reference(
                    ws, min_row=2, min_col=1, max_row=ws.max_row, max_col=1
                )
                self.setup_excel_chart(
                    values=values,
                    categories=categories,
                    chart_height=chart_height,
                    chart_width=chart_width,
                    chart_yaxis_titles=chart_yaxis_titles[i],
                    chart_yaxis_scaling_mins=chart_yaxis_scaling_mins[i],
                    chart_yaxis_scaling_maxes=chart_yaxis_scaling_maxes[i],
                    chart_yaxis_major_unit=chart_yaxis_major_unit[i],
                )

                ws.add_chart(self.chart, "B" + str(5 + 20 * i))
        wb.save(self.input_file.replace("csv", "xlsx"))

    def setup_excel_chart(
        self,
        values,
        categories,
        chart_height,
        chart_width,
        chart_yaxis_titles,
        chart_yaxis_scaling_mins,
        chart_yaxis_scaling_maxes,
        chart_yaxis_major_unit,
    ):
        """ excel chart setup

        """

        self.chart = LineChart()
        self.chart.add_data(values, titles_from_data=True)
        self.chart.set_categories(categories)
        self.chart.height = chart_height
        self.chart.width = chart_width
        # self.chart.x_axis.title = "abc"
        self.chart.y_axis.title = chart_yaxis_titles
        self.chart.y_axis.scaling.min = chart_yaxis_scaling_mins
        self.chart.y_axis.scaling.max = chart_yaxis_scaling_maxes
        self.chart.y_axis.majorUnit = chart_yaxis_major_unit
        self.chart.y_axis.numFmt = "0.0"
        self.chart.x_axis.tickLblPos = "low"

        # chart.y_axis.majorGridlines = None

        for i in range(len(self.chart.series)):
            series = self.chart.series[i]
            series.marker.symbol = "circle"
            # series.marker.size = 10
            series.graphicalProperties.line.noFill = True

    def make_graph(
        self,
        df_columns_list,
        figsize=(10, 5.5),
        file_name="default",
        fontsize=14,
        hlines="",
        legends=[],
        rotation=45,
        style=["bo", "ro", "go", "yo"],
        xlabel="",
        ylabel="",
        yticks=[],
    ):
        """ make specified graph from dataframe using matplotlib
            Args:
                df_columns_list (list): dataframe columns list to make graph
                fontsize (int): font size
                xlabel (str): xlabel
                ylabel (str): ylabel
                style (list): marker style
                file_name (str): filename
                rotation (int): rotation
                figsize (list): figure size
                yticks (list): yticks
                hlines (int): yaxis hline
                legends (list): legend

            Returns:
                None
        """
        global image_count

        if self.groupby:
            for name, group in self.data_df.groupby(self.groupby):

                # skip if dataframe has missing value
                if group[df_columns_list].isnull().values.sum() != 0:
                    continue

                num_of_index = len(group[df_columns_list].index)

                # adjust graph in case few index or missing value
                if (num_of_index == 3) & (
                    group[df_columns_list].isnull().values.sum() != 0
                ):
                    xmargin = 0.5
                else:
                    xmargin = 0.1

                self.setup_fig_and_ax(figsize, bottom=0.3, xmargin=xmargin)

                self.ax.set_xticks(
                    [i for i in range(group.shape[0])]
                )  # set number of label

                # df_tmp = group[df_columns_list].copy()
                # df_tmp = df_tmp.dropna(how="any", axis=0)

                # df_tmp.plot(
                #     ax=self.ax, ylim=ylim, style=style, legend=True, fontsize=fontsize,
                # )
                print(group[df_columns_list])
                group[df_columns_list].plot(
                    ax=self.ax,
                    ylim=yticks[:2],
                    style=style,
                    legend=True,
                    fontsize=fontsize,
                )

                self.adjust_graph_params(
                    rotation=rotation,
                    xlabel=xlabel,
                    ylabel=ylabel,
                    fontsize=fontsize,
                    yticks=yticks,
                    hlines=hlines,
                    num_of_index=num_of_index,
                    legends=legends,
                    group_name=name,
                )

                num = f"{image_count:03}_"
                file_name_full = (
                    self.folder_path + num + name + "_" + file_name + ".png"
                )
                plt.savefig(file_name_full)
                plt.close("all")

                self.add_slide_to_pptx(
                    title=num + name + "_" + file_name,
                    slide_count=self.slide_count,
                    layout=11,
                )

                self.add_picture_to_pptx(file_name_full=file_name_full)

        else:
            self.setup_fig_and_ax(figsize, bottom=0.3)

            # set number of label
            self.ax.set_xticks([i for i in range(self.data_df.shape[0])])

            self.data_df[df_columns_list].plot(
                ax=self.ax, ylim=yticks[:2], style=style, legend=True, fontsize=fontsize
            )

            self.adjust_graph_params(
                rotation=rotation,
                xlabel=xlabel,
                ylabel=ylabel,
                fontsize=fontsize,
                yticks=yticks,
                hlines=hlines,
                num_of_index=len(self.data_df.index),
                legends=legends,
            )

            num = f"{image_count:03}_"
            file_name_full = self.folder_path + num + file_name + ".png"
            plt.savefig(self.folder_path + num + file_name + ".png")
            plt.close("all")
            self.add_slide_to_pptx(
                title=num + file_name, slide_count=self.slide_count, layout=11,
            )

            self.add_picture_to_pptx(file_name_full=file_name_full)

    def make_vix_graph(
        self,
        file_name,
        nega_pin_file,
        posi_pin_file,
        fontsize=14,
        figsize=(10, 5.5),
        legends=[],
        rotation=0,
        xlabel="",
        ylabel="",
    ):
        """ make vix graph from posi/nega wave data file using matplotlib
            Args:
                file_name (str): file name
                nega_pin_file (str): csv nega pin file name
                posi_pin_file (str): csv posi pin file name
                fontsize (int): fontsize
                figsize (list): figure size
                legends (list): legend labels
                rotation (int): xlabel rotation value
                xlabel (str): xlabel
                ylabel (str): ylabel
            Returns:
                None
        """
        global image_count

        self.setup_fig_and_ax(figsize=figsize, xmargin=0.01)

        df_posi = pd.read_csv(posi_pin_file, header=None)
        df_nega = pd.read_csv(nega_pin_file, header=None)

        df_posi = df_posi.set_axis(["t", "wck_t"], axis=1)
        df_nega = df_nega.set_axis(["t", "wck_c"], axis=1)

        df_posi = df_posi.set_index("t")
        df_nega = df_nega.set_index("t")

        df_posi_nega = pd.concat([df_posi, df_nega], axis=1)

        # make diff column
        df_posi_nega["wck_t-wck_c"] = df_posi_nega["wck_t"] - df_posi_nega["wck_c"]

        # make dataframe df_vix. df_vix has 4 data which wck_t - wck_c is close to 0
        # close to 0 or 0 means cross point
        df_tmp = df_posi_nega.copy()
        df_vix = pd.DataFrame()
        for i in range(4):
            val = self.getNearestValue(df_tmp["wck_t-wck_c"].values.tolist(), 0)
            min_row1 = df_tmp[df_tmp["wck_t-wck_c"] == val]
            df_vix = pd.concat([df_vix, min_row1])
            df_tmp = df_tmp.drop(min_row1.index)

        # get average in case there is no cross point in data
        df_vix["(wck_t+wck_c)/2"] = (df_vix["wck_t"] + df_vix["wck_c"]) / 2

        df_vix = df_vix["(wck_t+wck_c)/2"]
        df_vix = df_vix.reset_index()
        df_vix_list = df_vix.values.tolist()

        df_posi_nega = df_posi_nega.drop("wck_t-wck_c", axis=1)
        df_posi_nega.plot(ax=self.ax)

        # reference level line
        self.ax.hlines(
            y=0,
            xmin=0,
            xmax=len(df_posi_nega.index) / 16,
            color="black",
            linestyle="dashed",
        )

        # add x, y cordinates to graph
        for df_vix_p in df_vix_list:
            x = df_vix_p[0]
            y = df_vix_p[1]
            self.ax.text(x + 0.2, y, f"({x:.2f}, {y:.2f})")
            self.ax.annotate(
                "", xy=[x, 0], xytext=df_vix_p, arrowprops=dict(arrowstyle="<->")
            )

        self.adjust_graph_params(
            rotation=rotation,
            xlabel=xlabel,
            ylabel=ylabel,
            fontsize=fontsize,
            yticks=[],
            hlines="",
            num_of_index=[],
            legends=legends,
        )
        num = f"{image_count:03}_"
        file_name_full = self.folder_path + num + file_name + ".png"
        plt.savefig(file_name_full)
        plt.close("all")
        self.add_slide_to_pptx(
            title=num + file_name, slide_count=self.slide_count, layout=11,
        )
        self.add_picture_to_pptx(file_name_full=file_name_full)

    def getNearestValue(self, list, num):
        """ return nearest value of num from list

        Args:
            list (list): list of num
            num (int): num of value

        Returns:
            nearest value of num from list
        """
        idx = np.abs(np.asarray(list) - num).argmin()
        return list[idx]

    def setup_fig_and_ax(self, figsize=(16, 9), bottom=0.2, xmargin=0.1):
        """

        Args:
            None

        Returns:
            None
        """
        self.fig = plt.figure(figsize=figsize)  # create figure object
        self.ax = self.fig.add_subplot(1, 1, 1, xmargin=xmargin)  # create axes object
        self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter("%.1f"))
        self.fig.subplots_adjust(bottom=bottom)

    def adjust_graph_params(
        self,
        fontsize=14,
        hlines="",
        legends=[],
        num_of_index=0,
        rotation=0,
        xlabel="",
        ylabel="",
        yticks=[],
        group_name="",
    ):
        """

        Args:
            None

        Returns:
            None
        """
        plt.xticks(rotation=rotation)
        self.ax.set_ylabel(ylabel, fontsize=fontsize)
        self.ax.set_xlabel(xlabel, fontsize=fontsize)
        self.ax.legend(labels=legends, fontsize=fontsize, loc="upper right")

        if yticks:
            self.ax.set_yticks(np.arange(yticks[0], yticks[1] + yticks[2], yticks[2]))

        # add limit line in case AT condition Vih/Vil=1V/0V
        match_at_condition = re.match(r".*Vih1r0V_Vil0r0V", group_name)
        if match_at_condition and hlines:
            self.ax.hlines(
                y=hlines,
                xmin=0,
                xmax=num_of_index - 1,
                linestyle={"dashed"},
                colors=["gray"],
            )

    def add_picture_to_pptx(self, file_name_full):
        """

        Args:
            None

        Returns:
            None
        """
        global image_count
        image = self.active_presentation.Slides(self.slide_count).Shapes.AddPicture(
            FileName=file_name_full, LinkToFile=-1, SaveWithDocument=-1, Left=0, Top=0,
        )
        image.Left = self.slide_width / 2 - image.Width / 2
        image.Top = self.slide_height / 2 - image.Height / 2
        image_count += 1

    def add_summary_table_to_pptx(
        self,
        title,
        cell_width=[],
        cell_height=20,
        items=[],
        groupby_table="",
        header_rename_dict={},
    ):
        """ add summary table to pptx

        add slide, add table to pptx.
            Args:
                title (str): slide title
                cell_width (list): cell width
                cell_height (list): cell height
                items (list): items for table
                groupby_table (str): group name of table in case separate table by group
                header_rename_dict (dict): specify header original and after name in case rename

            Returns:
                None
        """

        self.add_slide_to_pptx(title=title, slide_count=self.slide_count, layout=4)
        data_list_to_table_df = self.data_df.reset_index()
        self.add_table(
            df=data_list_to_table_df,
            items=items,
            cell_width=cell_width,
            cell_height=cell_height,
            slide_width=self.slide_width,
            slide_height=self.slide_height,
            header_rename_dict=header_rename_dict,
        )

        if groupby_table:
            for name, group in self.data_df.groupby(groupby_table):
                slide_count = self.active_presentation.Slides.Count
                self.add_slide_to_pptx(title=name, slide_count=slide_count, layout=4)
                data_list_to_table_df = group.reset_index()
                self.add_table(
                    df=data_list_to_table_df,
                    items=items,
                    cell_width=cell_width,
                    cell_height=cell_height,
                    slide_width=self.slide_width,
                    slide_height=self.slide_height,
                    header_rename_dict=header_rename_dict,
                )

    def add_slide_to_pptx(self, title, slide_count, layout):
        """

        Args:
            None

        Returns:
            None
        """
        self.slide = self.active_presentation.Slides.Add(
            Index=slide_count + 1, Layout=layout
        )
        self.slide.Select()
        self.slide.Shapes(1).TextFrame.TextRange.Text = title
        self.slide.Shapes(1).TextFrame.TextRange.Font.Size = 28
        self.slide_count += 1

    def add_table(
        self,
        df,
        items,
        cell_width,
        cell_height,
        slide_width,
        slide_height,
        header_rename_dict,
    ):
        """

        Args:
            None

        Returns:
            None
        """
        df = df.loc[:, items]
        print(df)
        data_list_to_table = df.values.tolist()
        data_list_to_table.insert(0, df.columns.tolist())

        table_rows = len(data_list_to_table)
        table_columns = len(data_list_to_table[0])
        table = self.slide.Shapes.AddTable(table_rows, table_columns).Table

        for i in range(table_rows):
            for j in range(table_columns):
                tr = table.Cell(i + 1, j + 1).Shape.TextFrame.TextRange
                try:
                    tr.Text = f"{data_list_to_table[i][j]:.1f}"
                except ValueError:
                    if header_rename_dict:
                        for key, value in header_rename_dict.items():
                            if key == data_list_to_table[i][j]:
                                tr.Text = value
                                break
                            else:
                                tr.Text = data_list_to_table[i][j]
                    else:
                        tr.Text = data_list_to_table[i][j]

                tr.Font.Size = 10

        for i in range(1, table.Columns.Count + 1):
            table.Columns(i).Width = cell_width[i - 1]

        for i in range(1, table.Rows.Count + 1):
            table.Rows(i).Height = cell_height

        # adjust table position
        shape = self.slide.Shapes(2)
        shape.Left = slide_width / 2 - shape.width / 2
        shape.Top = slide_height / 6

    def save_pptx(self, file_name):
        """save pptx file
            Args:
                file_name (str): file name

            Returns:
                None
        """
        self.active_presentation.SaveAs(
            FileName=os.getcwd() + "/" + str(date_now) + "_" + file_name
        )

    def mul3(self, x):
        return x * 1e3

    def mul12(self, x):
        return x * 1e12

    def mulm9(self, x):
        return x * 1e-9

    def adjust_unit(self):
        """

        Args:
            None

        Returns:
            None
        """
        if "Eheight" in self.data_df.columns:
            self.data_df["Eheight"] = self.data_df["Eheight"].apply(self.mul3)

        if "Ewidth" in self.data_df.columns:
            self.data_df["Ewidth"] = self.data_df["Ewidth"].apply(self.mul12)

        if "Risetime" in self.data_df.columns:
            self.data_df["Risetime"] = self.data_df["Risetime"].apply(self.mul12)

        if "Falltime" in self.data_df.columns:
            self.data_df["Falltime"] = self.data_df["Falltime"].apply(self.mul12)

        if "Frequency" in self.data_df.columns:
            self.data_df["Frequency"] = self.data_df["Frequency"].apply(self.mulm9)

        if "Vamplitude" in self.data_df.columns:
            self.data_df["Vamplitude"] = self.data_df["Vamplitude"].apply(self.mul3)

        if "Vpp" in self.data_df.columns:
            self.data_df["Vpp"] = self.data_df["Vpp"].apply(self.mul3)

        if "Vmaximum" in self.data_df.columns:
            self.data_df["Vmaximum"] = self.data_df["Vmaximum"].apply(self.mul3)

        if "Vminimum" in self.data_df.columns:
            self.data_df["Vminimum"] = self.data_df["Vminimum"].apply(self.mul3)


if __name__ == "__main__":

    CELL_WIDTH_BASE = 72
    DATA_START_COLUMNS = 9
    FOLDER_PATH = os.getcwd() + "/sample_log/"
    FILE_NAME = "8GPE_TEST.pptx"
    PE = "8GPE_"
    FREQ_YTICKS = [3.0, 5.0, 0.2]
    DUTY_YTICKS = [40.0, 60.0, 5]
    TRTF_YTICKS = [30.0, 70.0, 10]
    EHEIGHT_YTICS = [300, 400, 10]
    EWIDTH_YTICKS = [60, 120, 10]

    wave_data_overview = WaveData(
        file_name="result_overview3.csv",
        folder_path=FOLDER_PATH,
        groupby="Pkind_Vi",
        new_presentation=False,
        index="Pin_Rate",
    )
    wave_data_overview.make_vix_graph(
        posi_pin_file="./sample_log/posi.csv",
        nega_pin_file="./sample_log/nega.csv",
        file_name="Vix",
        ylabel="mV",
        legends=["wck_t", "wck_c"],
    )
    wave_data_overview.make_graph(
        df_columns_list=["Frequency"],
        file_name=PE + "Frequency",
        yticks=FREQ_YTICKS,
        ylabel="GHz",
        legends=["Freq(GHz)"],
    )
    wave_data_overview.make_graph(
        df_columns_list=["Dutycycle"],
        file_name=PE + "Duty",
        yticks=DUTY_YTICKS,
        ylabel="%",
        legends=["Duty(%)"],
    )
    wave_data_overview.make_graph(
        df_columns_list=["Risetime", "Falltime"],
        yticks=TRTF_YTICKS,
        file_name=PE + "Risetime_Falltime",
        ylabel="ps",
        hlines=60,
        legends=["Tr(ps)", "Tf(ps)"],
    )
    wave_data_overview.add_summary_table_to_pptx(
        title="overview",
        cell_width=[
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 1.1,
            CELL_WIDTH_BASE * 1.1,
        ],
        items=["Pin", "Vi", "Rate", "Risetime", "Falltime"],
        groupby_table="Vi",
        header_rename_dict={"Risetime": "Risetime(ps)", "Falltime": "Falltime(ps)"},
    )
    wave_data_overview.make_excel_graphs(
        data_start_column=DATA_START_COLUMNS,
        chart_yaxis_titles=["ps", "ps", "%", "GHz", "mV", "mV", "mV", "mV"],
        chart_yaxis_scaling_mins=[0, 0, 45, 3.9, 400, 400, 400, -60],
        chart_yaxis_scaling_maxes=[100, 100, 55, 4.1, 600, 600, 600, 60],
        chart_yaxis_major_unit=[20, 20, 2, 0.05, 20, 20, 20, 20],
    )
    wave_data_eye = WaveData(
        file_name="result_eye.csv",
        folder_path=FOLDER_PATH,
        groupby="Pkind_Vi",
        new_presentation=False,
    )
    wave_data_eye.make_graph(
        df_columns_list=["Eheight"],
        yticks=EHEIGHT_YTICS,
        file_name=PE + "Eheight",
        ylabel="mV",
        legends=["Eye Height(mV)"],
    )
    wave_data_eye.make_graph(
        df_columns_list=["Ewidth"],
        yticks=EWIDTH_YTICKS,
        file_name=PE + "Ewidth",
        legends=["Eye Width(ps)"],
        ylabel="ps",
    )
    wave_data_eye.add_summary_table_to_pptx(
        title="eye",
        cell_width=[
            # CELL_WIDTH_BASE * 5,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 2,
            CELL_WIDTH_BASE * 2,
        ],
        items=["Pin", "Vi", "Rate", "Eheight", "Ewidth"],
        groupby_table="Vi",
    )
    wave_data_eye.make_excel_graphs(
        data_start_column=DATA_START_COLUMNS,
        chart_yaxis_titles=["ps", "mV"],
        chart_yaxis_scaling_mins=[300, 0],
        chart_yaxis_scaling_maxes=[400, 10],
        chart_yaxis_major_unit=[20, 2],
    )
    wave_data_eye.make_excel_graph(
        data_start_column=DATA_START_COLUMNS,
        data_end_column=8,
        chart_yaxis_title="ps",
        chart_yaxis_scaling_min=0,
        chart_yaxis_scaling_max=400,
        chart_yaxis_major_unit=50,
        chart_position="L5",
    )
    wave_data_overview.save_pptx(file_name=FILE_NAME)
