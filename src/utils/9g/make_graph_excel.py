from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference


wb = load_workbook("./sample_log/test.xlsx")
# ws = wb["Sheet1"]
ws = wb.worksheets[0]

data_start_column = 2
# chart_positions = ["B5", "B25", "B45"]
chart_yaxis_title = ["ps", "mV"]
chart_yaxis_scaling_min = [300, 0]
chart_yaxis_scaling_max = [400, 10]
for i in range(ws.max_column - 1):
    values = Reference(
        ws,
        min_row=1,
        min_col=data_start_column + i,
        max_row=ws.max_row,
        max_col=data_start_column + i,
    )
    categories = Reference(ws, min_row=2, min_col=1, max_row=ws.max_row, max_col=1)

    chart = LineChart()
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 16
    chart.x_axis.title = ""
    chart.y_axis.title = chart_yaxis_title[i]
    chart.y_axis.scaling.min = chart_yaxis_scaling_min[i]
    chart.y_axis.scaling.max = chart_yaxis_scaling_max[i]

    print(chart.series.length)
    series = chart.series[0]
    series.marker.symbol = "circle"
    # series.marker.size = 10
    series.graphicalProperties.line.noFill = True

    ws.add_chart(chart, "B" + str(5 + 20 * i))
wb.save("abc.xlsx")
