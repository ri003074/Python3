import csv
import datetime
import matplotlib.pyplot as plt
import os
import pandas as pd
import win32com.client
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
import re

# from glob import glob
# import sys

now = datetime.datetime.now()
date_now = now.strftime("%Y%m%d%H%M")

"""
P773A1_eye_Vih1r000V_Vil0r000V_Vt0r500V_Rate0r250ns_Duty0r500	 EHEIGHT	3.64E-01	 EWIDTH	9.62E-11
P773A1_overview_Vih1r000V_Vil0r000V_Vt0r500V_Rate0r250ns_Duty0r500	 RISETIME	4.76E-11	 FALLTIME	4.54E-11	 DUTYCYCLE	4.97E+01	 FREQUENCY	4.00E+09	 VAMPLITUDE	5.05E-01	 VPP	5.43E-01	 VMAXIMUM	5.21E-01	 VMINIMUM	-2.23E-02
"""


class WaveData:
    def __init__(self, filename, folderpath, header=[], groupby="", index="Pin_Rate"):
        self.file_name = filename
        self.folder_path = folderpath
        self.input_file = self.folder_path + self.file_name
        self.data_df = ""
        self.data_list = ""
        self.header = header
        self.groupby = groupby
        self.index = index

    def make_df_and_xlsx(self):
        with open(self.input_file, mode="r", encoding="utf-8-sig") as csvfile:
            reader = csv.reader(csvfile)
            data = []

            for rows in reader:
                match = re.match(r"(.*?)_.*?_(.*?_.*?_.*?)_(.*?)_", rows[0])
                if match:  # add extra information for data sort
                    rows.insert(0, "Condition")
                    rows.insert(2, "Pin")
                    rows.insert(3, match.group(1))
                    rows.insert(4, "Vi")
                    rows.insert(5, match.group(2))
                    rows.insert(6, "Rate")
                    rows.insert(
                        7, match.group(3).replace("Rate0r", "").replace("ns", "ps")
                    )

                dic = OrderedDict()
                for i in range(0, len(rows) - 1, 2):
                    try:
                        dic[rows[i].replace(" ", "").capitalize()] = float(
                            rows[i + 1].replace(" ", "")
                        )
                    except ValueError:
                        dic[rows[i].replace(" ", "").capitalize()] = rows[
                            i + 1
                        ].replace(" ", "")

                data.append(dic)

            self.data_df = pd.DataFrame(data)
            self.data_df.insert(  # add extra information for data sort
                1, "Pin_Rate", self.data_df["Pin"] + "_" + self.data_df["Rate"]
            )
            self.data_df.insert(
                2, "Pin_Vi", self.data_df["Pin"] + "_" + self.data_df["Vi"]
            )

            self.adjust_unit()  # adjust unit of dataframe

            if self.header:
                self.data_df = self.data_df.set_axis(self.header, axis="columns")

            self.data_df = self.data_df.set_index(self.index)

            self.data_list = self.data_df.reset_index().values.tolist()
            self.data_list.insert(0, self.data_df.reset_index().columns.to_list())

            with pd.ExcelWriter(self.input_file.replace("csv", "xlsx")) as writer:
                self.data_df.to_excel(writer, sheet_name="summary")

                if self.groupby:
                    for name, group in self.data_df.groupby(self.groupby):
                        group.to_excel(writer, sheet_name=name)

    def make_excel_graph(
        self,
        data_start_column,
        data_end_column,
        chart_yaxis_title="",
        chart_yaxis_scaling_min="",
        chart_yaxis_scaling_max="",
        chart_yaxis_major_unit="",
        chart_position="E5",
        chart_height=9,
        chart_width=16,
    ):
        wb = load_workbook(self.input_file.replace("csv", "xlsx"))
        for i in range(len(wb.worksheets)):
            ws = wb.worksheets[i]

            values = Reference(
                ws,
                min_row=1,
                min_col=data_start_column,
                max_row=ws.max_row,
                max_col=data_end_column,
            )
            categories = Reference(
                ws, min_row=2, min_col=1, max_row=ws.max_row, max_col=1
            )

            self.setup_excel_chart(
                values,
                categories,
                chart_height,
                chart_width,
                chart_yaxis_titles=chart_yaxis_title,
                chart_yaxis_scaling_mins=chart_yaxis_scaling_min,
                chart_yaxis_scaling_maxes=chart_yaxis_scaling_max,
                chart_yaxis_major_unit=chart_yaxis_major_unit,
            )

            ws.add_chart(self.chart, chart_position)
        wb.save(self.input_file.replace("csv", "xlsx"))

    def make_excel_graphs(
        self,
        data_start_column,
        chart_yaxis_titles=[],
        chart_yaxis_scaling_mins=[],
        chart_yaxis_scaling_maxes=[],
        chart_yaxis_major_unit=[],
        chart_height=9,
        chart_width=16,
    ):
        wb = load_workbook(self.input_file.replace("csv", "xlsx"))
        for i in range(len(wb.worksheets)):
            ws = wb.worksheets[i]

            for i in range(ws.max_column + data_start_column * -1 + 1):
                values = Reference(
                    ws,
                    min_row=1,
                    min_col=data_start_column + i,
                    max_row=ws.max_row,
                    max_col=data_start_column + i,
                )
                categories = Reference(
                    ws, min_row=2, min_col=1, max_row=ws.max_row, max_col=1
                )
                self.setup_excel_chart(
                    values,
                    categories,
                    chart_height,
                    chart_width,
                    chart_yaxis_titles=chart_yaxis_titles[i],
                    chart_yaxis_scaling_mins=chart_yaxis_scaling_mins[i],
                    chart_yaxis_scaling_maxes=chart_yaxis_scaling_maxes[i],
                    chart_yaxis_major_unit=chart_yaxis_major_unit[i],
                )

                ws.add_chart(self.chart, "B" + str(5 + 20 * i))
        wb.save(self.input_file.replace("csv", "xlsx"))

    def setup_excel_chart(
        self,
        values,
        categories,
        chart_height,
        chart_width,
        chart_yaxis_titles,
        chart_yaxis_scaling_mins,
        chart_yaxis_scaling_maxes,
        chart_yaxis_major_unit,
    ):
        self.chart = LineChart()
        self.chart.add_data(values, titles_from_data=True)
        self.chart.set_categories(categories)
        self.chart.height = chart_height
        self.chart.width = chart_width
        # self.chart.x_axis.title = "abc"
        self.chart.y_axis.title = chart_yaxis_titles
        self.chart.y_axis.scaling.min = chart_yaxis_scaling_mins
        self.chart.y_axis.scaling.max = chart_yaxis_scaling_maxes
        self.chart.y_axis.majorUnit = chart_yaxis_major_unit
        self.chart.y_axis.numFmt = "0.0"
        self.chart.x_axis.tickLblPos = "low"

        # chart.y_axis.majorGridlines = None

        for i in range(len(self.chart.series)):
            series = self.chart.series[i]
            series.marker.symbol = "circle"
            # series.marker.size = 10
            series.graphicalProperties.line.noFill = True

    def make_graph(
        self,
        df_columns_list,
        ylim,
        fontsize=14,
        xlabel="",
        ylabel="",
        style=["bo", "yo", "ro", "go"],
        filename="",
        rotation=0,
        figsize=(16, 9),
    ):

        self.setup_fig_and_ax(figsize)

        self.ax.set_xticks(
            [i for i in range(self.data_df.shape[0])]
        )  # set number of label

        self.data_df[df_columns_list].plot(
            ax=self.ax, ylim=ylim, style=style, legend=True, fontsize=fontsize
        )

        self.adjust_graph_params(rotation, xlabel, ylabel, fontsize)

        plt.savefig(self.folder_path + filename + ".png")
        plt.close("all")

        if self.groupby:
            for name, group in self.data_df.groupby(self.groupby):
                self.setup_fig_and_ax(figsize)

                self.ax.set_xticks(
                    [i for i in range(group.shape[0])]
                )  # set number of label
                group[df_columns_list].plot(
                    ax=self.ax, ylim=ylim, style=style, legend=True, fontsize=fontsize
                )

                self.adjust_graph_params(rotation, xlabel, ylabel, fontsize)

                plt.savefig(self.folder_path + name + "_" + filename + ".png")
                plt.close("all")

    def adjust_graph_params(self, rotation, xlabel, ylabel, fontsize):
        plt.xticks(rotation=rotation)
        self.ax.set_ylabel(ylabel, fontsize=fontsize)
        self.ax.set_xlabel(xlabel, fontsize=fontsize)
        self.ax.legend(fontsize=fontsize)

    def setup_fig_and_ax(self, figsize):
        self.fig = plt.figure(figsize=figsize)  # create figure object
        self.ax = self.fig.add_subplot(1, 1, 1)  # create axes object
        self.ax.yaxis.set_major_formatter(plt.FormatStrFormatter("%.1f"))
        # self.fig.subplots_adjust(bottom=0.2)

    def add_table_to_pptx(self, new_presentation, title, cell_width, items=[]):
        if new_presentation:
            pptx = win32com.client.Dispatch("PowerPoint.Application")
            pptx.Visible = True
            self.active_presentation = pptx.Presentations.Open(
                os.getcwd() + "/advtemplate_mini.pptx"
            )
        else:
            pptx = win32com.client.GetActiveObject("PowerPoint.Application")
            self.active_presentation = pptx.ActivePresentation

        slide_width = self.active_presentation.PageSetup.SlideWidth
        slide_height = self.active_presentation.PageSetup.SlideHeight
        slide_count = self.active_presentation.Slides.Count

        slide = self.active_presentation.Slides.Add(Index=slide_count + 1, Layout=4)
        slide.Select()
        slide.Shapes(1).TextFrame.TextRange.Text = title

        data_list_to_table = self.data_list.copy()
        if items:
            items_removed = list(set(data_list_to_table[0]) ^ set(items))
            items_removed_index = []
            for item in items_removed:
                items_removed_index.append(data_list_to_table[0].index(item))

            data_list_to_table = []
            for item in self.data_list:
                for i in sorted(items_removed_index, reverse=True):
                    item.pop(i)
                data_list_to_table.append(item)

        table_rows = len(data_list_to_table)
        table_columns = len(data_list_to_table[0])
        table = slide.Shapes.AddTable(table_rows, table_columns).Table

        for i in range(table_rows):
            for j in range(table_columns):
                tr = table.Cell(i + 1, j + 1).Shape.TextFrame.TextRange
                try:
                    tr.Text = f"{data_list_to_table[i][j]:.1f}"
                except ValueError:
                    tr.Text = data_list_to_table[i][j]

                tr.Font.Size = 14

        for i in range(1, table.Columns.Count + 1):
            table.Columns(i).Width = cell_width[i - 1]

        shape = slide.Shapes(2)
        shape.Left = slide_width / 2 - shape.width / 2
        shape.Top = slide_height / 6

    def save_pptx(self, file_name):
        self.active_presentation.SaveAs(
            FileName=os.getcwd() + "/" + str(date_now) + "_" + file_name
        )

    def mul3(self, x):
        return x * 1e3

    def mul12(self, x):
        return x * 1e12

    def mulm9(self, x):
        return x * 1e-9

    def adjust_unit(self):
        if "Eheight" in self.data_df.columns:
            self.data_df["Eheight"] = self.data_df["Eheight"].apply(self.mul3)

        if "Ewidth" in self.data_df.columns:
            self.data_df["Ewidth"] = self.data_df["Ewidth"].apply(self.mul12)

        if "Risetime" in self.data_df.columns:
            self.data_df["Risetime"] = self.data_df["Risetime"].apply(self.mul12)

        if "Falltime" in self.data_df.columns:
            self.data_df["Falltime"] = self.data_df["Falltime"].apply(self.mul12)

        if "Frequency" in self.data_df.columns:
            self.data_df["Frequency"] = self.data_df["Frequency"].apply(self.mulm9)

        if "Vamplitude" in self.data_df.columns:
            self.data_df["Vamplitude"] = self.data_df["Vamplitude"].apply(self.mul3)

        if "Vpp" in self.data_df.columns:
            self.data_df["Vpp"] = self.data_df["Vpp"].apply(self.mul3)

        if "Vmaximum" in self.data_df.columns:
            self.data_df["Vmaximum"] = self.data_df["Vmaximum"].apply(self.mul3)

        if "Vminimum" in self.data_df.columns:
            self.data_df["Vminimum"] = self.data_df["Vminimum"].apply(self.mul3)


CELL_WIDTH_BASE = 72
DATA_START_COLUMNS = 7


data_summarize_eye = WaveData(
    filename="result_eye.csv",
    folderpath=os.getcwd() + "/sample_log/",
    header=[  # header order is important. must be same as data after extra items added
        "Condition",
        "Pin_Rate",
        "Pin_Vi",
        "Pin",
        "Vi",
        "Rate",
        "Eye Height(mV)",
        "Eye Width(mV)",
    ],
    groupby="Vi",
)
data_summarize_eye.make_df_and_xlsx()
data_summarize_eye.make_graph(
    df_columns_list=["Eye Height(mV)"],
    ylim=[300, 400],
    filename="8g_eheight",
    ylabel="ps",
)
data_summarize_eye.make_graph(
    df_columns_list=["Eye Width(mV)"], ylim=[0, 10], filename="8g_ewidth",
)
data_summarize_eye.make_graph(
    df_columns_list=["Eye Width(mV)", "Eye Height(mV)"],
    ylim=[0, 500],
    filename="8g_ewidth_eheight",
)
data_summarize_eye.add_table_to_pptx(
    new_presentation=False,
    title="eye",
    cell_width=[
        # CELL_WIDTH_BASE * 5,
        CELL_WIDTH_BASE * 2,
        CELL_WIDTH_BASE * 2,
        CELL_WIDTH_BASE * 2,
        CELL_WIDTH_BASE * 2,
        CELL_WIDTH_BASE * 2,
        CELL_WIDTH_BASE * 2,
        CELL_WIDTH_BASE * 2,
    ],
    items=["Condition", "Pin", "Vi", "Rate", "Eye Height(mV)", "Eye Width(mV)"],
)
data_summarize_eye.make_excel_graphs(
    data_start_column=DATA_START_COLUMNS,
    chart_yaxis_titles=["ps", "mV"],
    chart_yaxis_scaling_mins=[300, 0],
    chart_yaxis_scaling_maxes=[400, 10],
    chart_yaxis_major_unit=[20, 2],
)
data_summarize_eye.make_excel_graph(
    data_start_column=DATA_START_COLUMNS,
    data_end_column=8,
    chart_yaxis_title="ps",
    chart_yaxis_scaling_min=0,
    chart_yaxis_scaling_max=400,
    chart_yaxis_major_unit=50,
    chart_position="L5",
)
data_summarize_overview = WaveData(
    filename="result_overview2.csv",
    folderpath=os.getcwd() + "/sample_log/",
    groupby="Vi",
)
data_summarize_overview.make_df_and_xlsx()
data_summarize_overview.make_graph(
    df_columns_list=["Risetime"], ylim=[40, 70], filename="8g_risetime",
)
data_summarize_overview.make_graph(
    df_columns_list=["Falltime"], ylim=[40, 70], filename="8g_falltime",
)
data_summarize_overview.make_graph(
    df_columns_list=["Falltime", "Risetime"],
    ylim=[40, 70],
    filename="8g_falltime_risetime",
)
data_summarize_overview.add_table_to_pptx(
    new_presentation=False,
    title="overview",
    cell_width=[
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 4,
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 1.1,
        CELL_WIDTH_BASE * 1.1,
    ],
    items=["Pin", "Vi", "Rate", "Risetime", "Falltime"],
)
data_summarize_overview.make_excel_graphs(
    data_start_column=DATA_START_COLUMNS,
    chart_yaxis_titles=["ps", "ps", "%", "GHz", "mV", "mV", "mV", "mV"],
    chart_yaxis_scaling_mins=[0, 0, 45, 3.9, 400, 400, 400, -50],
    chart_yaxis_scaling_maxes=[100, 100, 55, 4.1, 600, 600, 600, 50],
    chart_yaxis_major_unit=[20, 20, 2, 0.05, 20, 20, 20, 20],
)
data_summarize_eye.save_pptx(file_name="test")
