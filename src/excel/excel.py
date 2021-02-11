from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

wb = Workbook()
ws = wb.active

data = [
    ["Apple", 1],
    ["Banana", 2],
    ["Grap", 3],
]

for row_index in range(len(data)):
    for column_index in range(len(data[row_index])):
        ws.cell(row=row_index + 1, column=column_index + 1).value = data[row_index][
            column_index
        ]

# ws.cell(row=1, column=1).value = "Apple"
# ws.cell(row=2, column=1).value = "Banana"
# ws.cell(row=3, column=1).value = "Grape"

# ws.cell(row=1, column=2).value = 1
# ws.cell(row=2, column=2).value = 2
# ws.cell(row=3, column=2).value = 3

# wb.save("output_tmp.xlsx")

values = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=3)
categories = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=3)

chart = LineChart()
chart.legend = None
chart.title = "Fruits"

chart.add_data(values)
chart.set_categories(categories)

ws.add_chart(chart, "B4")
wb.save("output.xlsx")
