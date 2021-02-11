import wx
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


class MainFrame(wx.Frame):
    def __init__(self, *args, **kw):
        super(MainFrame, self).__init__(*args, **kw)

        self.wb = Workbook()
        self.ws = self.wb.active
        self.data = []
        self.open_dialog()
        self.read_file()
        self.show_data()
        self.make_excel()
        self.make_graph()

    def open_dialog(self):
        filter = "*.csv;*.txt"
        dialog = wx.FileDialog(None, "select file", style=wx.FD_OPEN, wildcard=filter)
        dialog.ShowModal()
        self.input_file_path = dialog.GetPath()

    def read_file(self):
        with open(self.input_file_path, "r") as f:
            for line in f.read().splitlines():
                self.data.append(line.split(","))

    def make_excel(self):
        for row_index in range(len(self.data)):
            for column_index in range(len(self.data[row_index])):
                try:
                    self.ws.cell(
                        row=row_index + 1, column=column_index + 1
                    ).value = float(self.data[row_index][column_index])

                except ValueError:
                    self.ws.cell(
                        row=row_index + 1, column=column_index + 1
                    ).value = self.data[row_index][column_index]

    def make_graph(self):
        values = Reference(self.ws, min_col=2, min_row=1, max_col=3, max_row=4)
        categories = Reference(self.ws, min_col=1, min_row=2, max_col=1, max_row=4)
        chart = LineChart()
        # chart.legend = True
        chart.title = "DC TEST RESULT"
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)

        self.ws.add_chart(chart, "B4")
        self.wb.save("sample.xlsx")

    def show_data(self):
        print(self.data)

    # def init_ui(self):
    #     self.SetTitle("dclog converter")
    #     self.SetSize((100, 100))
    #     self.Show()

    #     # panel = wx.Panel(self, -1, pos=(50, 50), size=(300, 200))
    #     panel = wx.Panel(self, -1, size=(100, 100), style=wx.TE_CENTER)
    #     sizer = wx.BoxSizer()
    #     sizer.AddStretchSpacer(1)
    #     sizer.Add(panel, 0, wx.ALIGN_CENTER_VERTICAL)
    #     sizer.AddStretchSpacer(1)
    #     self.SetSizer(sizer)

    #     # self.label = wx.StaticText(panel, -1, "", pos=(10, 10))

    #     # self.box = wx.TextCtrl(panel, -1, pos=(10, 50))

    #     # btn = wx.Button(panel, -1, "select file", pos=(10, 90))
    #     btn = wx.Button(panel, -1, "select file")
    #     btn.Bind(wx.EVT_BUTTON, self.clicked)

    # def clicked(self, event):
    #     # text = self.box.GetValue()
    #     # self.box.Clear()
    #     # self.label.SetLabel(text)
    #     filter = "*.csv;*.txt"
    #     dialog = wx.FileDialog(None, "select file", style=wx.FD_OPEN, wildcard=filter)
    #     dialog.ShowModal()
    #     print(dialog.GetPath())


if __name__ == "__main__":
    app = wx.App()
    frame = MainFrame(None)
    # frame.Show()
    # app.MainLoop()
    app.Destroy()
